import pandas as pd
from models.contractor import ScoreResult
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

COLOR_HEADER  = "1A365D"   # dark navy
COLOR_SUBHDR  = "2B6CB0"   # mid blue
COLOR_GREEN   = "C6EFCE"   # light green fill
COLOR_YELLOW  = "FFEB9C"   # light yellow fill
COLOR_RED     = "FFC7CE"   # light red fill
COLOR_GREEN_F = "1E7E34"
COLOR_YELLOW_F = "9C6500"
COLOR_RED_F   = "9C0006"
COLOR_ALT     = "F0F4F8"   # alternating row bg


def _header_font(bold=True, color="FFFFFF", size=10):
    return Font(bold=bold, color=color, size=size)


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _border():
    thin = Side(style="thin", color="D0D7E0")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def strip_polish_chars(text: str) -> str:
    """Replaces Polish characters with their ASCII equivalents to keep export file content standardized."""
    if not isinstance(text, str):
        return text
    mapping = {
        'ą': 'a', 'Ą': 'A',
        'ć': 'c', 'Ć': 'C',
        'ę': 'e', 'Ę': 'E',
        'ł': 'l', 'Ł': 'L',
        'ń': 'n', 'Ń': 'N',
        'ó': 'o', 'Ó': 'O',
        'ś': 's', 'Ś': 'S',
        'ź': 'z', 'Ź': 'Z',
        'ż': 'z', 'Ż': 'Z'
    }
    for pol, eng in mapping.items():
        text = text.replace(pol, eng)
    return text

def _score_colors(color_code: str):
    """Returns (fill_hex, font_hex) for a score cell based on risk color code."""
    if color_code == "green":
        return COLOR_GREEN, COLOR_GREEN_F
    if color_code == "red":
        return COLOR_RED, COLOR_RED_F
    return COLOR_YELLOW, COLOR_YELLOW_F


def _style_summary_sheet(ws, num_rows: int) -> None:
    """Apply formatting to the Summary sheet."""
    col_widths = [16, 32, 10, 10, 30, 30, 12, 16, 24, 22, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Header row
    for cell in ws[1]:
        cell.font = _header_font(bold=True, color="FFFFFF", size=10)
        cell.fill = _fill(COLOR_HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border()
    ws.row_dimensions[1].height = 30

    # Data rows
    for row_idx in range(2, num_rows + 2):
        is_alt = (row_idx % 2 == 0)
        for cell in ws[row_idx]:
            cell.border = _border()
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if is_alt:
                cell.fill = _fill("EDF2F7")


def _style_detail_sheet(ws, score: int, color_code: str, justification_start: int, num_rows: int) -> None:
    """Apply formatting to a per-NIP detail sheet."""
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 52

    fill_bg, font_color = _score_colors(color_code)

    for row_idx in range(1, num_rows + 1):
        row = ws[row_idx]
        label_cell = row[0]
        value_cell = row[1] if len(row) > 1 else None

        label_val = str(label_cell.value or "")

        # Section headers
        if label_val.isupper() and label_val not in ("NIP", "TAK", "NIE"):
            label_cell.font = _header_font(bold=True, color="FFFFFF", size=10)
            label_cell.fill = _fill(COLOR_SUBHDR)
            if value_cell:
                value_cell.font = _header_font(bold=True, color="FFFFFF", size=10)
                value_cell.fill = _fill(COLOR_SUBHDR)
        # Score row
        elif "WYNIK" in label_val:
            for cell in row:
                cell.fill = _fill(fill_bg)
                cell.font = Font(bold=True, color=font_color, size=11)
        # Rekomendacja row
        elif "REKOMENDACJA" in label_val:
            for cell in row:
                cell.fill = _fill(fill_bg)
                cell.font = Font(bold=True, color=font_color, size=10)
        # Justification items
        elif row_idx >= justification_start:
            if value_cell:
                value_cell.font = Font(color="4A5568", size=9, italic=True)
                value_cell.fill = _fill("FAFAFA")
            label_cell.font = Font(color="2B6CB0", bold=True)
            label_cell.fill = _fill("FAFAFA")
        else:
            # Alternating rows
            if row_idx % 2 == 0:
                for cell in row:
                    if cell.fill.fgColor.rgb in ("00000000", "FFFFFFFF", ""):
                        cell.fill = _fill("EDF2F7")

        for cell in row:
            cell.border = _border()
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[row_idx].height = 18


def export_results(results: list, path: str) -> None:
    """Export results to Excel — styled Summary sheet + one detail sheet per NIP."""
    import openpyxl

    summary_rows = []
    for r in results:
        is_cdata = hasattr(r, 'scoring') and isinstance(r.scoring, dict)
        nip = getattr(r, 'nip', '---')

        if is_cdata:
            score = r.scoring.get('total_score', 0)
            max_score = 100
            rec = r.scoring.get('risk_level', 'NIEZNANY')
            color_code = r.scoring.get('color_code', 'yellow')
            justifications = r.scoring.get('justifications', [])
        else:
            score = getattr(r, 'total', 0)
            max_score = 40
            rec = getattr(r, 'risk_level', None)
            rec = rec.value if rec else 'NIEZNANY'
            color_code = 'yellow'
            justifications = getattr(r, 'details', [])

        website = getattr(r, 'website_url', None) or "NIE ODNALEZIONO"
        ssl = "TAK" if getattr(r, 'ssl_valid', False) else "NIE"
        age = getattr(r, 'domain_age_days', None)
        domain_age = f"{age} dni (~{round(age/365,1)} lat)" if age is not None else "BRAK DANYCH"
        posts = getattr(r, 'days_since_last_post', None)
        activity = f"Ostatni wpis {posts} dni temu" if posts is not None else "BRAK DANYCH"
        nip_match = "TAK" if getattr(r, 'website_nip_matched', False) else "NIE"
        cap = getattr(r, 'share_capital', None)
        capital = f"{int(cap):,} PLN".replace(",", " ") if cap is not None else "BRAK DANYCH"
        bailiff = getattr(r, 'has_bailiff_proceedings', None)
        bailiff_str = "TAK" if bailiff is True else ("NIE" if bailiff is False else "BRAK DANYCH")
        just_short = "; ".join(justifications[:3]) if justifications else ""

        summary_rows.append({
            "NIP": nip,
            "Nazwa firmy": strip_polish_chars(getattr(r, 'legal_name', '---') if is_cdata else '---'),
            "Wynik": f"{score}/{max_score}",
            "Ocena ryzyka": strip_polish_chars(rec),
            "Uzasadnienie (skrot)": strip_polish_chars(just_short),
            "Status prawny": strip_polish_chars(getattr(r, 'status_prawny', '---') if is_cdata else '---'),
            "Status VAT": strip_polish_chars(getattr(r, 'status_vat', '---') if is_cdata else '---'),
            "Kapital": strip_polish_chars(capital),
            "Postep. komornicze": strip_polish_chars(bailiff_str),
            "Strona WWW": strip_polish_chars(website),
            "SSL": strip_polish_chars(ssl),
        })

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        pd.DataFrame(summary_rows).to_excel(writer, sheet_name="Podsumowanie", index=False)

        for r in results:
            is_cdata = hasattr(r, 'scoring') and isinstance(r.scoring, dict)
            nip = getattr(r, 'nip', '---')
            sheet_name = nip[:31]

            if is_cdata:
                score = r.scoring.get('total_score', 0)
                max_score = 100
                rec = r.scoring.get('risk_level', 'NIEZNANY')
                color_code = r.scoring.get('color_code', 'yellow')
                justifications = r.scoring.get('justifications', [])
            else:
                score = getattr(r, 'total', 0)
                max_score = 40
                rec = getattr(r, 'risk_level', None)
                rec = rec.value if rec else 'NIEZNANY'
                color_code = 'yellow'
                justifications = getattr(r, 'details', [])

            detail_rows = []

            def add(label, value):
                detail_rows.append({
                    "Pole": strip_polish_chars(label),
                    "Wartosc": strip_polish_chars(str(value)) if value is not None else "BRAK DANYCH"
                })

            add("DANE REJESTROWE", "")
            add("NIP", nip)
            if is_cdata:
                add("Nazwa firmy", getattr(r, 'legal_name', '---'))
                add("Status prawny", getattr(r, 'status_prawny', 'NIEZNANY'))
                add("Status VAT", getattr(r, 'status_vat', 'NIEZNANY'))
                add("Biala lista MF", "TAK" if getattr(r, 'rachunek_na_bialej_liscie', False) else "NIE")
                cap = getattr(r, 'share_capital', None)
                add("Kapital zakladowy", f"{int(cap):,} PLN".replace(",", " ") if cap is not None else None)
                bailiff = getattr(r, 'has_bailiff_proceedings', None)
                add("Postep. komornicze", "TAK" if bailiff is True else ("NIE" if bailiff is False else None))
                sanctions = getattr(r, 'on_sanctions_list', None)
                add("Na liscie sankcji", "TAK" if sanctions is True else ("NIE" if sanctions is False else None))

            add("WIARYGODNOSC CYFROWA", "")
            add("Strona internetowa", getattr(r, 'website_url', None) or "NIE ODNALEZIONO")
            add("Szyfrowanie SSL/TLS", "TAK" if getattr(r, 'ssl_valid', False) else "NIE")
            age = getattr(r, 'domain_age_days', None)
            add("Wiek domeny", f"{age} dni (~{round(age/365,1)} lat)" if age is not None else None)
            posts = getattr(r, 'days_since_last_post', None)
            add("Ostatnia aktywnosc (RSS)", f"Ostatni wpis {posts} dni temu" if posts is not None else None)
            add("Zgodnosc NIP na stronie", "TAK" if getattr(r, 'website_nip_matched', False) else "NIE")

            add("WYNIK KONCOWY", f"{score}/{max_score} pkt")
            add("REKOMENDACJA RYZYKA", rec)

            just_start = len(detail_rows) + 1
            add("UZASADNIENIE SZCZEGOLOWE", "")
            for j in justifications:
                detail_rows.append({"Pole": "  •", "Wartosc": strip_polish_chars(j)})

            pd.DataFrame(detail_rows).to_excel(writer, sheet_name=sheet_name, index=False)

        ws_book = writer.book

        # Style Summary sheet
        ws_sum = ws_book["Podsumowanie"]
        _style_summary_sheet(ws_sum, len(summary_rows))
        # Color-code score cells per row based on risk
        score_col_idx = 3  # "Wynik" column (1-indexed)
        for row_idx, r in enumerate(results, start=2):
            is_cdata = hasattr(r, 'scoring') and isinstance(r.scoring, dict)
            cc = r.scoring.get('color_code', 'yellow') if is_cdata else 'yellow'
            fill_bg, font_col = _score_colors(cc)
            cell = ws_sum.cell(row=row_idx, column=score_col_idx)
            cell.fill = _fill(fill_bg)
            cell.font = Font(bold=True, color=font_col, size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Style per-NIP sheets
        for r in results:
            is_cdata = hasattr(r, 'scoring') and isinstance(r.scoring, dict)
            nip = getattr(r, 'nip', '---')
            sheet_name = nip[:31]
            cc = r.scoring.get('color_code', 'yellow') if is_cdata else 'yellow'
            sc = r.scoring.get('total_score', 0) if is_cdata else 0
            ws_nip = ws_book[sheet_name]

            # Find justification start row (header row + data rows)
            num_rows = ws_nip.max_row
            just_row = num_rows  # fallback
            for ri in range(1, num_rows + 1):
                v = str(ws_nip.cell(row=ri, column=1).value or "")
                if "UZASADNIENIE" in v:
                    just_row = ri + 1
                    break

            _style_detail_sheet(ws_nip, sc, cc, just_row, num_rows)

def read_nips_from_excel(path: str) -> list[str]:
    """Read a list of NIPs from an Excel file (.xlsx). Looks for a column named 'NIP' or takes the first column."""
    try:
        df = pd.read_excel(path, engine="openpyxl")
        nip_col = next((col for col in df.columns if str(col).strip().upper() == "NIP"), None)
        
        if nip_col:
            nips = df[nip_col].dropna().astype(str).tolist()
        else:
            nips = df.iloc[:, 0].dropna().astype(str).tolist()
            
        cleaned_nips = []
        import re
        for nip in nips:
            clean = str(nip).replace(" ", "").replace("-", "").upper()
            if clean.endswith(".0"):
                clean = clean[:-2]
            
            if re.match(r"^[A-Z]{2}[A-Z0-9]{2,14}$", clean) or (clean.isdigit() and len(clean) == 10):
                cleaned_nips.append(clean)
        return cleaned_nips
    except Exception as e:
        print(f"Błąd podczas czytania pliku Excel: {e}")
        return []
